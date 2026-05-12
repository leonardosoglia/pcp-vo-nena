"""
exportar.py — gera XLSX a partir de uma folha salva no banco.

Estrutura do arquivo gerado:
    Sheet 1: Resumo            — KPIs do dia + cabeçalho
    Sheet 2: Cocada            — todos os quadros de cocada (Embalados, Cortados ①②③, Corte, Produção, Embalagem)
    Sheet 3: Palha             — quadros de palha
    Sheet 4: Papelzinho Joel   — 5 colunas × 6 sabores
    Sheet 5: PM/Balas/Obs      — produtos independentes + orientações

Uso:
    from exportar import gerar_xlsx_folha
    buf = gerar_xlsx_folha("2026-05-08")  # bytes prontos pra st.download_button
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import (
    get_folha_cocada, get_folha_palha, get_papelzinho_joel,
    get_pm_balas_doces, get_metas_45g, get_metas_mini_pet, get_pvirar_ideal,
    calcular_cortados, calcular_viradas_pvirar,
    SABORES_COCADA, SABORES_PALHA, SIGLA_COCADA, SIGLA_PALHA,
)

DIAS_PT = {0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta",
           4: "Sexta", 5: "Sábado", 6: "Domingo"}
DIAS_COL_METAS = {0: "segunda", 1: "terca", 2: "quarta", 3: "quinta", 4: "sexta"}

# Estilos Vó Nena
COR_TITULO = "C05621"
COR_DESTAQUE = "7B341E"
COR_SUBTITULO = "F7EDE2"
COR_LINHA_ALT = "FFFEFB"

FONT_TITULO_DOC = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_TITULO_BLOCO = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="7B341E")
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color="1A1A1A")

FILL_TITULO_DOC = PatternFill("solid", fgColor=COR_TITULO)
FILL_TITULO_BLOCO = PatternFill("solid", fgColor=COR_DESTAQUE)
FILL_SUBTITULO = PatternFill("solid", fgColor=COR_SUBTITULO)
FILL_LINHA_ALT = PatternFill("solid", fgColor=COR_LINHA_ALT)

THIN = Side(border_style="thin", color="D5C5B0")
BORDA = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _titulo_doc(ws, texto, ncols=6):
    """Linha de título grande (mesclada) no topo de cada sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=texto)
    cell.font = FONT_TITULO_DOC
    cell.fill = FILL_TITULO_DOC
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 32


def _titulo_bloco(ws, row, texto, ncols=6):
    """Cabeçalho de uma seção (mesclado)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=texto)
    cell.font = FONT_TITULO_BLOCO
    cell.fill = FILL_TITULO_BLOCO
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 22


def _escrever_tabela(ws, row_ini, col_ini, headers, data_rows, larguras=None):
    """Escreve cabeçalho + linhas de dados, com bordas e cor alternada.

    Retorna a próxima linha livre (após a tabela).
    """
    # Cabeçalho
    for j, h in enumerate(headers):
        cell = ws.cell(row=row_ini, column=col_ini + j, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_SUBTITULO
        cell.alignment = CENTER
        cell.border = BORDA

    # Dados
    for i, row in enumerate(data_rows):
        for j, val in enumerate(row):
            cell = ws.cell(row=row_ini + 1 + i, column=col_ini + j, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = CENTER if j > 0 else LEFT
            cell.border = BORDA
            if i % 2 == 1:
                cell.fill = FILL_LINHA_ALT
            # Coluna 0 (sabor) em negrito
            if j == 0:
                cell.font = FONT_LABEL

    # Larguras
    if larguras:
        for j, w in enumerate(larguras):
            ws.column_dimensions[get_column_letter(col_ini + j)].width = w

    return row_ini + 1 + len(data_rows)


# ══════════════════════════════════════════════════════════════════════════════
# SHEETS
# ══════════════════════════════════════════════════════════════════════════════
def _build_sheet_resumo(wb, data, dia_pt, fc_dict, fp_dict, pj_dict, pbd):
    ws = wb.create_sheet("Resumo")
    _titulo_doc(ws, f"PCP Vó Nena · Folha de {data} ({dia_pt})", ncols=4)

    row = 3
    _titulo_bloco(ws, row, "🎯 KPIs do dia", ncols=4); row += 1

    def _add(label, valor):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row, column=1).border = BORDA
        ws.cell(row=row, column=2, value=valor).font = FONT_NORMAL
        ws.cell(row=row, column=2).border = BORDA
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

    total_emb_45 = sum(int(r.get("emb_45g") or 0) for r in fc_dict.values())
    total_emb_mi = sum(int(r.get("emb_mini") or 0) for r in fc_dict.values())
    total_emb_pet = sum(int(r.get("emb_pet") or 0) for r in fc_dict.values())
    total_cort_45 = sum(int(r.get("cort1_45g") or 0) for r in fc_dict.values())
    total_band_prod = sum(int(r.get("ord_prod_band") or 0) for r in fc_dict.values())
    total_emb_ord = sum(int(r.get("ord_emb_45g") or 0) + int(r.get("ord_emb_mini") or 0) for r in fc_dict.values())

    _add("Embalados 45g (total)", total_emb_45)
    _add("Embalados Mini (total)", total_emb_mi)
    _add("Embalados Pet (total)", total_emb_pet)
    _add("Cortados ① 45g (total)", total_cort_45)
    _add("Bandejas a produzir (total)", total_band_prod)
    _add("Unidades a embalar (total)", total_emb_ord)
    if pbd:
        _add("PM (cnt)", int(pbd.get("cnt_pm") or 0))
        _add("Balas (cnt)", int(pbd.get("cnt_balas") or 0))
        _add("Doces (und)", int(pbd.get("cnt_doces_displays") or 0))
        _add("Ordem PM", int(pbd.get("ord_pm") or 0))
        _add("Ordem Balas (tachos)", int(pbd.get("ord_balas") or 0))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    row += 2
    cell = ws.cell(row=row, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    cell.font = Font(name="Calibri", size=9, italic=True, color="888888")


def _build_sheet_cocada(wb, data, fc_dict, pj_dict):
    ws = wb.create_sheet("Cocada")
    _titulo_doc(ws, f"Cocada — {data}", ncols=8)
    row = 3

    # EMBALADOS
    _titulo_bloco(ws, row, "🎁 Embalados", ncols=6); row += 1
    headers = ["Sabor", "45g (und)", "Mini (und)", "Pet (und)", "Potes 260g", "Potes 605g"]
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        rows.append([
            s,
            int(r.get("emb_45g") or 0),
            int(r.get("emb_mini") or 0),
            int(r.get("emb_pet") or 0),
            int(r.get("emb_potes_260g") or 0),
            int(r.get("emb_potes_605g") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 12, 12, 12, 12, 12])
    row += 2

    # CORTADOS ① ② ③
    _titulo_bloco(ws, row, "✂️ Cortados — ① ② ③ (em unidades)", ncols=8); row += 1
    headers = ["Sabor", "① 45g", "② 45g", "③ 45g", "① Mini", "② Mini", "③ Mini", "① Pet"]
    cort_calc = {r["sabor"]: r for r in calcular_cortados(data)}
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        c = cort_calc.get(s, {})
        rows.append([
            s,
            int(r.get("cort1_45g") or 0),
            c.get("c2_45g") or 0,
            c.get("c3_45g") if c.get("c3_45g") is not None else "—",
            int(r.get("cort1_mini") or 0),
            c.get("c2_mini") or 0,
            c.get("c3_mini") if c.get("c3_mini") is not None else "—",
            int(r.get("cort1_pet") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20] + [10] * 7)
    row += 2

    # CORTE DE COCADA
    _titulo_bloco(ws, row, "🔪 Corte de Cocada — Ordens (bandejas)", ncols=4); row += 1
    headers = ["Sabor", "45g (band)", "Mini (band)", "Pet (band)"]
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        rows.append([
            s,
            int(r.get("ord_corte_45g") or 0),
            int(r.get("ord_corte_mini") or 0),
            int(r.get("ord_corte_pet") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 14, 14, 14])
    row += 2

    # PARÂMETRO REAL
    _titulo_bloco(ws, row, "⚙️ Parâmetro Real do dia (45g · Mini · Pet)", ncols=4); row += 1
    headers = ["Sabor", "Param 45g", "Param Mini", "Param Pet"]
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        rows.append([
            s,
            int(r.get("param_real_45g") or 0),
            int(r.get("param_real_mini") or 0),
            int(r.get("param_real_pet") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 14, 14, 14])
    row += 2

    # PRODUÇÃO
    _titulo_bloco(ws, row, "🏭 Produção — Ordens", ncols=5); row += 1
    headers = ["Sabor", "Bandejas", "Virada", "Potes 260g", "Potes 605g"]
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        rows.append([
            s,
            int(r.get("ord_prod_band") or 0),
            int(r.get("ord_prod_virada") or 0),
            int(r.get("ord_prod_potes_260g") or 0),
            int(r.get("ord_prod_potes_605g") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 12, 12, 14, 14])
    row += 2

    # EMBALAGEM
    _titulo_bloco(ws, row, "📦 Embalagem — Ordens (unidades)", ncols=3); row += 1
    headers = ["Sabor", "45g (und)", "Mini (und)"]
    rows = []
    for s in SABORES_COCADA:
        r = fc_dict.get(s, {})
        rows.append([
            s,
            int(r.get("ord_emb_45g") or 0),
            int(r.get("ord_emb_mini") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 14, 14])

    # VIRADAS / P/VIRAR (derivado)
    row += 2
    _titulo_bloco(ws, row, "🔄 Viradas e P/Virar (derivado)", ncols=6); row += 1
    headers = ["Sabor", "Viradas ①", "Viradas ②", "P/Virar ①", "P/Virar ②", "Meta"]
    vp_calc = {r["sabor"]: r for r in calcular_viradas_pvirar(data)}
    rows = []
    for s in SABORES_COCADA:
        v = vp_calc.get(s, {})
        rows.append([
            s,
            int(v.get("vir1") or 0),
            int(v.get("vir2") or 0),
            int(v.get("pv1") or 0),
            int(v.get("pv2") or 0),
            int(v.get("pv_meta") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[20, 12, 12, 12, 12, 10])


def _build_sheet_palha(wb, data, fp_dict):
    ws = wb.create_sheet("Palha")
    _titulo_doc(ws, f"Palha — {data}", ncols=4)
    row = 3

    # EMBALADOS PALHA
    _titulo_bloco(ws, row, "🌾 Embalados Palha", ncols=3); row += 1
    headers = ["Sabor", "50g (und)", "Pet 160g (und)"]
    rows = []
    for s in SABORES_PALHA:
        r = fp_dict.get(s, {})
        rows.append([
            f"{s} ({SIGLA_PALHA.get(s, '')})",
            int(r.get("emb_50g") or 0),
            int(r.get("emb_pet") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[26, 14, 14])
    row += 2

    # COLUNA PALHA
    _titulo_bloco(ws, row, "🌾 Coluna PALHA — Bandejas (Leonardo)", ncols=3); row += 1
    headers = ["Sabor", "Bandejas (band)", "Pós-corte (band)"]
    rows = []
    for s in SABORES_PALHA:
        r = fp_dict.get(s, {})
        rows.append([
            f"{s} ({SIGLA_PALHA.get(s, '')})",
            int(r.get("cont_band_palha") or 0),
            int(r.get("cont_band_pos_corte") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[26, 16, 16])
    row += 2

    # PRODUÇÃO PALHA
    _titulo_bloco(ws, row, "🏭 Produção Palha — Ordens", ncols=2); row += 1
    headers = ["Sabor", "Bandejas"]
    rows = []
    for s in SABORES_PALHA:
        r = fp_dict.get(s, {})
        rows.append([
            f"{s} ({SIGLA_PALHA.get(s, '')})",
            int(r.get("ord_prod_band") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[26, 14])
    row += 2

    # CORTE PALHA
    _titulo_bloco(ws, row, "🔪 Corte Palha — Ordens", ncols=3); row += 1
    headers = ["Sabor", "50g (band)", "Pet (band)"]
    rows = []
    for s in SABORES_PALHA:
        r = fp_dict.get(s, {})
        rows.append([
            f"{s} ({SIGLA_PALHA.get(s, '')})",
            int(r.get("ord_corte_50g") or 0),
            int(r.get("ord_corte_pet") or 0),
        ])
    row = _escrever_tabela(ws, row, 1, headers, rows, larguras=[26, 14, 14])


def _build_sheet_papelzinho(wb, data, pj_dict):
    ws = wb.create_sheet("Papelzinho Joel")
    _titulo_doc(ws, f"Papelzinho do Joel — {data}", ncols=6)
    row = 3
    _titulo_bloco(ws, row, "📝 5 colunas × 6 sabores", ncols=6); row += 1

    headers = ["Sabor", "45g (und)", "30g (und)", "P (und)", "PV (band)", "V (band)"]
    rows = []
    for s in SABORES_COCADA:
        r = pj_dict.get(s, {})
        sigla = SIGLA_COCADA.get(s, s[:1])
        rows.append([
            f"{s} ({sigla})",
            int(r.get("joel_45g") or 0) if s != "ZERO" else "—",
            int(r.get("joel_mini") or 0),
            int(r.get("joel_pet") or 0),
            int(r.get("joel_pv") or 0),
            int(r.get("joel_v") or 0),
        ])
    _escrever_tabela(ws, row, 1, headers, rows, larguras=[24, 12, 12, 12, 12, 12])


def _build_sheet_pmbd_obs(wb, data, pbd):
    ws = wb.create_sheet("PM Balas Obs")
    _titulo_doc(ws, f"PM · Balas · Doces · Orientações — {data}", ncols=4)
    row = 3

    if pbd:
        _titulo_bloco(ws, row, "🍞 Contagens (topo da folha)", ncols=2); row += 1
        for label, key in [("PM (cnt)", "cnt_pm"), ("Balas (cnt)", "cnt_balas"), ("Doces (und)", "cnt_doces_displays")]:
            ws.cell(row=row, column=1, value=label).font = FONT_LABEL
            ws.cell(row=row, column=1).border = BORDA
            ws.cell(row=row, column=2, value=int(pbd.get(key) or 0))
            ws.cell(row=row, column=2).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = BORDA
            row += 1

        row += 2
        _titulo_bloco(ws, row, "📋 Ordens do dia (linha 36)", ncols=2); row += 1
        for label, key in [("Ordem PM", "ord_pm"), ("Ordem Balas (tachos)", "ord_balas")]:
            ws.cell(row=row, column=1, value=label).font = FONT_LABEL
            ws.cell(row=row, column=1).border = BORDA
            ws.cell(row=row, column=2, value=int(pbd.get(key) or 0))
            ws.cell(row=row, column=2).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = BORDA
            row += 1

        if pbd.get("ord_amanha_obs"):
            row += 2
            _titulo_bloco(ws, row, "📅 PM amanhã", ncols=4); row += 1
            ws.cell(row=row, column=1, value=pbd.get("ord_amanha_obs", "")).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = LEFT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

        # Orientações (campo único)
        obs_completo = (pbd.get("obs", "") or "").strip()
        for legado_campo, label in [("obs_joel", "Joel"), ("obs_gil", "Gil"), ("obs_leonilia", "Leonília")]:
            extra = (pbd.get(legado_campo, "") or "").strip()
            if extra:
                obs_completo += ("\n" if obs_completo else "") + f"[{label}]: {extra}"

        if obs_completo:
            row += 2
            _titulo_bloco(ws, row, "📋 Orientações", ncols=4); row += 1
            ws.cell(row=row, column=1, value=obs_completo).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = LEFT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.row_dimensions[row].height = 80
    else:
        ws.cell(row=row, column=1, value="(Sem dados de PM/Balas/Doces para essa data.)").font = FONT_NORMAL

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


# ══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO PÚBLICA
# ══════════════════════════════════════════════════════════════════════════════
def gerar_xlsx_folha(data: str) -> bytes:
    """Gera arquivo XLSX da folha do dia. Retorna bytes prontos pra download."""
    fc_dict = {r["sabor"]: r for r in get_folha_cocada(data)}
    fp_dict = {r["sabor"]: r for r in get_folha_palha(data)}
    pj_dict = {r["sabor"]: r for r in get_papelzinho_joel(data)}
    pbd = get_pm_balas_doces(data)

    try:
        d_obj = datetime.strptime(data, "%Y-%m-%d").date()
        dia_pt = DIAS_PT.get(d_obj.weekday(), "")
    except Exception:
        dia_pt = ""

    wb = Workbook()
    # Remove a sheet padrão "Sheet"
    wb.remove(wb.active)

    _build_sheet_resumo(wb, data, dia_pt, fc_dict, fp_dict, pj_dict, pbd)
    _build_sheet_cocada(wb, data, fc_dict, pj_dict)
    _build_sheet_palha(wb, data, fp_dict)
    _build_sheet_papelzinho(wb, data, pj_dict)
    _build_sheet_pmbd_obs(wb, data, pbd)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
